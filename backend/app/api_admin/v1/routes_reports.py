# app/api_admin/v1/routes_reports.py
from __future__ import annotations
from typing import Optional, List, Any, cast, Dict
from datetime import datetime, timedelta
import logging
import csv
import io

from fastapi import APIRouter, Depends, Query, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, cast as sa_cast, Numeric
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError

from app.core.deps import get_db, require_admin
from app.db import models  # Order, OrderItem, Product, Artist

# Router bez /api/admin/v1 (prefix se doplní v app.main)
router = APIRouter(prefix="/reports", tags=["admin: reports"])

log = logging.getLogger("app.reports")

ALLOWED_STATUSES = {"draft", "pending_payment", "paid", "shipped", "canceled"}


def _parse_status_list(s: Optional[str]) -> Optional[List[str]]:
    if not s:
        return None
    vals = [v.strip() for v in s.split(",") if v.strip()]
    bad = [v for v in vals if v not in ALLOWED_STATUSES]
    if bad:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Neplatný status: {', '.join(bad)}",
        )
    return vals or None


def _pick_date_col(O) -> Optional[Any]:
    # prefer created_at, fallback updated_at
    if hasattr(O, "created_at"):
        return getattr(O, "created_at")
    if hasattr(O, "updated_at"):
        return getattr(O, "updated_at")
    return None


def _base_query(
    db: Session,
    status_filter: Optional[str],
    since: Optional[datetime],
    until: Optional[datetime],
):
    O = models.Order
    OI = models.OrderItem
    P = models.Product
    A = models.Artist

    qty_num = sa_cast(func.coalesce(OI.qty, 0), Numeric)
    price_num = sa_cast(func.coalesce(OI.unit_price, 0), Numeric)
    qty_expr = func.sum(qty_num).label("qty_sold")
    revenue_expr = func.sum(qty_num * price_num).label("revenue")

    q = (
        db.query(
            OI.product_id.label("product_id"),
            func.coalesce(P.title, "").label("title"),
            func.coalesce(A.name, "").label("artist_name"),
            qty_expr,
            revenue_expr,
        )
        .join(O, OI.order_id == O.id)
        .outerjoin(P, OI.product_id == P.id)
        .outerjoin(A, P.artist_id == A.id)
    )

    statuses = _parse_status_list(status_filter) or ["paid"]
    q = q.filter(O.status.in_(statuses))

    has_created_at = hasattr(O, "created_at")
    date_col = getattr(O, "created_at", None) if has_created_at else getattr(O, "updated_at", None)
    if date_col is not None:
        if since is not None:
            q = q.filter(date_col >= since)
        if until is not None:
            q = q.filter(date_col < until)

    q = q.group_by(OI.product_id, P.title, A.name)
    return q, has_created_at, statuses


# ---------- JSON ----------
@router.get("/sold-products", dependencies=[Depends(require_admin)])
def report_sold_products(
    status_filter: Optional[str] = Query("paid"),
    since: Optional[datetime] = Query(None),
    until: Optional[datetime] = Query(None),
    sort: str = Query("qty", pattern="^(qty|revenue)$"),
    order: str = Query("desc", pattern="^(asc|desc)$"),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    try:
        q, has_created_at, statuses = _base_query(db, status_filter, since, until)
        sub = q.subquery("sold_products")

        total = db.query(func.count()).select_from(sub).scalar() or 0

        sort_col = sub.c.revenue if sort == "revenue" else sub.c.qty_sold
        sort_expr = sort_col.desc() if order.lower() == "desc" else sort_col.asc()

        rows = (
            db.query(sub.c.product_id, sub.c.title, sub.c.artist_name, sub.c.qty_sold, sub.c.revenue)
            .order_by(sort_expr)
            .offset(offset)
            .limit(limit)
            .all()
        )

        items = [
            {
                "product_id": int(r.product_id) if r.product_id is not None else None,
                "title": r.title or "",
                "artist_name": r.artist_name or "",
                "qty_sold": int(r.qty_sold or 0),
                "revenue": float(r.revenue or 0),
            }
            for r in rows
        ]

        return {
            "total": int(total),
            "limit": limit,
            "offset": offset,
            "items": items,
            "filters": {
                "status": statuses or ["paid"],
                "since": since.isoformat() if since else None,
                "until": until.isoformat() if until else None,
                "sort": sort,
                "order": order,
            },
            "has_created_at_filter": has_created_at,
        }

    except SQLAlchemyError as e:
        log.exception("sold-products: SQLAlchemyError")
        raise HTTPException(status_code=500, detail=f"SQLAlchemyError: {e}") from e


# ---------- CSV ----------
@router.get("/sold-products.csv", dependencies=[Depends(require_admin)])
def report_sold_products_csv(
    status_filter: Optional[str] = Query("paid"),
    since: Optional[datetime] = Query(None),
    until: Optional[datetime] = Query(None),
    sort: str = Query("qty", pattern="^(qty|revenue)$"),
    order: str = Query("desc", pattern="^(asc|desc)$"),
    db: Session = Depends(get_db),
):
    try:
        q, _, _ = _base_query(db, status_filter, since, until)
        sub = q.subquery("sold_products")
        sort_col = sub.c.revenue if sort == "revenue" else sub.c.qty_sold
        sort_expr = sort_col.desc() if order.lower() == "desc" else sort_col.asc()

        rows = (
            db.query(sub.c.product_id, sub.c.title, sub.c.artist_name, sub.c.qty_sold, sub.c.revenue)
            .order_by(sort_expr)
            .all()
        )

        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["product_id", "title", "artist_name", "qty_sold", "revenue"])
        for r in rows:
            w.writerow(
                [
                    int(r.product_id) if r.product_id is not None else "",
                    r.title or "",
                    r.artist_name or "",
                    int(r.qty_sold or 0),
                    float(r.revenue or 0),
                ]
            )
        buf.seek(0)

        filename = (
            f'sold_products_{(status_filter or "paid").replace(",","_")}_'
            f'{(since.isoformat() if since else "any")}_'
            f'{(until.isoformat() if until else "any")}_{sort}_{order}.csv'
        )
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except SQLAlchemyError as e:
        log.exception("sold-products.csv: SQLAlchemyError")
        raise HTTPException(status_code=500, detail=f"SQLAlchemyError: {e}") from e


# ---------- XLSX ----------
@router.get("/sold-products.xlsx", dependencies=[Depends(require_admin)])
def report_sold_products_xlsx(
    status_filter: Optional[str] = Query("paid"),
    since: Optional[datetime] = Query(None),
    until: Optional[datetime] = Query(None),
    sort: str = Query("qty", pattern="^(qty|revenue)$"),
    order: str = Query("desc", pattern="^(asc|desc)$"),
    db: Session = Depends(get_db),
):
    try:
        # import uvnitř, aby openpyxl nebyl tvrdý runtime požadavek pro všechny
        from openpyxl import Workbook
        from openpyxl.worksheet.worksheet import Worksheet
        from openpyxl.utils import get_column_letter

        q, _, _ = _base_query(db, status_filter, since, until)
        sub = q.subquery("sold_products")
        sort_col = sub.c.revenue if sort == "revenue" else sub.c.qty_sold
        sort_expr = sort_col.desc() if order.lower() == "desc" else sort_col.asc()

        rows = (
            db.query(sub.c.product_id, sub.c.title, sub.c.artist_name, sub.c.qty_sold, sub.c.revenue)
            .order_by(sort_expr)
            .all()
        )

        wb = Workbook()
        ws: Worksheet = cast(Worksheet, wb.active)
        ws.title = "Sold products"
        ws.append(["product_id", "title", "artist_name", "qty_sold", "revenue"])
        for r in rows:
            ws.append(
                [
                    int(r.product_id) if r.product_id is not None else None,
                    r.title or "",
                    r.artist_name or "",
                    int(r.qty_sold or 0),
                    float(r.revenue or 0),
                ]
            )

        # auto-width
        for col in range(1, ws.max_column + 1):
            length = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                length = max(length, len(str(val)) if val is not None else 0)
            ws.column_dimensions[get_column_letter(col)].width = min(60, max(10, length + 2))

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        filename = (
            f'sold_products_{(status_filter or "paid").replace(",","_")}_'
            f'{(since.isoformat() if since else "any")}_'
            f'{(until.isoformat() if until else "any")}_{sort}_{order}.xlsx'
        )
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Chybí balíček 'openpyxl'. Nainstaluj: pip install openpyxl",
        )
    except SQLAlchemyError as e:
        log.exception("sold-products.xlsx: SQLAlchemyError")
        raise HTTPException(status_code=500, detail=f"SQLAlchemyError: {e}") from e


# ---------- ORDERS TIMESERIES ----------
@router.get("/sold-products/{product_id}/orders", dependencies=[Depends(require_admin)])
def report_product_orders(
    product_id: int,
    since: Optional[datetime] = Query(None),
    until: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Vrací seznam objednávek (paid) pro daný produkt.
    """
    try:
        O = models.Order
        OI = models.OrderItem
        P = models.Product

        # Základní query: najdi objednávky s tímto produktem ve statusu paid
        q = db.query(O).join(OI, OI.order_id == O.id).filter(
            OI.product_id == product_id,
            O.status == "paid"
        )

        # Filtruj podle datumu
        date_col = None
        if hasattr(O, "created_at"):
            date_col = getattr(O, "created_at")
        elif hasattr(O, "updated_at"):
            date_col = getattr(O, "updated_at")

        if date_col is not None:
            if since is not None:
                q = q.filter(date_col >= since)
            if until is not None:
                q = q.filter(date_col < until)

        orders = q.all()

        # Pro každou objednávku načti items
        result_orders = []
        for order in orders:
            items = db.query(OI).filter(OI.order_id == order.id).all()

            items_data = []
            for item in items:
                product = db.query(P).filter(P.id == item.product_id).first()
                items_data.append({
                    "product_id": int(item.product_id) if item.product_id is not None else None,
                    "title": product.title if product else "(bez názvu)",
                    "qty": int(item.qty or 1),
                    "unit_price": float(item.unit_price or 0),
                })

            result_orders.append({
                "id": int(order.id),
                "total": float(order.total or 0),
                "status": order.status,
                "created_at": order.created_at.isoformat() if hasattr(order, "created_at") and order.created_at else None,
                "payment_method": getattr(order, "payment_method", None),
                "shipping_method": getattr(order, "shipping_method", None),
                "items": items_data,
            })

        return {
            "product_id": product_id,
            "orders": result_orders,
        }

    except SQLAlchemyError as e:
        log.exception(f"sold-products/{product_id}/orders: SQLAlchemyError")
        raise HTTPException(status_code=500, detail=f"SQLAlchemyError: {e}") from e


# ---------- ORDERS TIMESERIES ----------
@router.get("/orders-timeseries", dependencies=[Depends(require_admin)])
def report_orders_timeseries(
    status_filter: Optional[str] = Query("paid", description="CSV statusů (např. paid,shipped)"),
    since: Optional[datetime] = Query(None, description="Od (včetně)"),
    until: Optional[datetime] = Query(None, description="Do (exkluzivně)"),
    bucket: str = Query("day", pattern="^(day|month)$"),
    db: Session = Depends(get_db),
):
    """
    Vrací časové řady počtu objednávek a tržeb.
    """
    O = models.Order

    # vyber datumový sloupec
    date_attr = None
    if hasattr(O, "created_at"):
        date_attr = getattr(O, "created_at")
    elif hasattr(O, "updated_at"):
        date_attr = getattr(O, "updated_at")

    try:
        q = db.query(O)

        # statusy
        statuses = _parse_status_list(status_filter) or ["paid"]
        q = q.filter(O.status.in_(statuses))

        # čas
        if date_attr is not None:
            if since is not None:
                q = q.filter(date_attr >= since)
            if until is not None:
                q = q.filter(date_attr < until)

        orders: List[models.Order] = q.all()

        # rozsah, když since/until chybí
        if since is None or until is None:
            if orders:
                all_dates = []
                for o in orders:
                    d = getattr(o, "created_at", None) or getattr(o, "updated_at", None)
                    if isinstance(d, datetime):
                        all_dates.append(d)
                if all_dates:
                    min_d = min(all_dates)
                    max_d = max(all_dates)
                else:
                    now = datetime.utcnow()
                    min_d = now
                    max_d = now
            else:
                now = datetime.utcnow()
                min_d = now
                max_d = now
            if since is None:
                since = datetime(min_d.year, min_d.month, 1)
            if until is None:
                if bucket == "month":
                    if min_d.month == 12:
                        until = datetime(min_d.year + 1, 1, 1)
                    else:
                        until = datetime(min_d.year, min_d.month + 1, 1)
                else:
                    until = since + timedelta(days=31)

        def day_key(d: datetime) -> str:
            return f"{d.year:04d}-{d.month:02d}-{d.day:02d}T00:00:00"

        def month_key(d: datetime) -> str:
            return f"{d.year:04d}-{d.month:02d}-01T00:00:00"

        make_key = month_key if bucket == "month" else day_key

        # skeleton
        series_keys: List[str] = []
        if bucket == "month":
            cur_y, cur_m = since.year, since.month
            end_y, end_m = until.year, until.month
            while (cur_y, cur_m) <= (end_y, end_m):
                k = month_key(datetime(cur_y, cur_m, 1))
                series_keys.append(k)
                if cur_m == 12:
                    cur_y += 1
                    cur_m = 1
                else:
                    cur_m += 1
        else:
            cur = since
            while cur <= until:
                series_keys.append(day_key(cur))
                cur = cur + timedelta(days=1)

        rev_map: Dict[str, float] = {k: 0.0 for k in series_keys}
        ord_map: Dict[str, int] = {k: 0 for k in series_keys}

        for o in orders:
            d = getattr(o, "created_at", None) or getattr(o, "updated_at", None)
            if not isinstance(d, datetime):
                continue
            k = make_key(d)
            if k not in rev_map:
                continue
            try:
                rev_map[k] += float(getattr(o, "total", 0) or 0)
            except Exception:
                pass
            ord_map[k] += 1

        revenue_series = [{"date": k, "value": float(v)} for k, v in rev_map.items()]
        orders_series = [{"date": k, "value": int(v)} for k, v in ord_map.items()]

        totals_revenue = float(sum(v for v in rev_map.values()))
        totals_orders = int(sum(v for v in ord_map.values()))

        return {
            "series": {"revenue": revenue_series, "orders": orders_series},
            "totals": {"revenue": totals_revenue, "orders": totals_orders},
            "filters": {
                "status": statuses,
                "since": since.isoformat() if since else None,
                "until": until.isoformat() if until else None,
                "bucket": bucket,
            },
        }

    except SQLAlchemyError as e:
        log.exception("orders-timeseries: SQLAlchemyError")
        raise HTTPException(status_code=500, detail=f"SQLAlchemyError: {e}") from e
